import os
import csv
import re
from datetime import datetime
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
from zzzzz import main as run_latency_analysis

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "uploads"
REPORTS_FOLDER = "reports"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


def read_cids_from_csv(filepath):
    """Extract conversation IDs from CSV file"""
    cids = []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            col_names = reader.fieldnames or []
            cid_col = next(
                (col for col in col_names if col.lower() in ["cid", "conversation_id", "call_id", "convo_id"]),
                col_names[0] if col_names else None
            )

            if not cid_col:
                return None, "No valid CID column found in CSV"

            for row in reader:
                cid = row.get(cid_col, "").strip()
                while cid and cid[0] in '"\',' + ';':
                    cid = cid[1:]
                while cid and cid[-1] in '"\',' + ';':
                    cid = cid[:-1]
                cid = cid.strip()
                if cid:
                    cids.append(cid)

        return cids, None
    except Exception as e:
        return None, f"Error reading CSV: {str(e)}"


def read_cids_from_txt(filepath):
    """Extract conversation IDs from TXT file (one ID per line)"""
    cids = []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                cid = line.strip()
                if cid and not cid.startswith('#'):
                    cids.append(cid)

        if not cids:
            return None, "No conversation IDs found in TXT file"
        return cids, None
    except Exception as e:
        return None, f"Error reading TXT: {str(e)}"


def read_cids_from_xlsx(filepath):
    """Extract conversation IDs from XLSX file"""
    cids = []
    try:
        if not load_workbook:
            return None, "openpyxl not installed. Install with: pip install openpyxl"

        workbook = load_workbook(filepath)
        worksheet = workbook.active

        cid_col_idx = None
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value and str(cell.value).lower() in ["cid", "conversation_id", "call_id", "convo_id"]:
                cid_col_idx = col_idx
                break

        if cid_col_idx is None:
            cid_col_idx = 1

        for row_idx, row in enumerate(worksheet.iter_rows(min_col=cid_col_idx, max_col=cid_col_idx, values_only=True), 1):
            if row_idx == 1:
                continue
            cid = row[0]
            if cid:
                cids.append(str(cid).strip())

        if not cids:
            return None, "No conversation IDs found in XLSX file"
        return cids, None
    except Exception as e:
        return None, f"Error reading XLSX: {str(e)}"


def read_cids_from_file(filepath):
    """Extract conversation IDs from CSV, TXT, or XLSX file"""
    file_ext = filepath.lower().split('.')[-1]

    try:
        if file_ext == 'csv':
            return read_cids_from_csv(filepath)
        elif file_ext == 'txt':
            return read_cids_from_txt(filepath)
        elif file_ext in ['xlsx', 'xls']:
            return read_cids_from_xlsx(filepath)
        else:
            return None, f"Unsupported file format: {file_ext}"
    except Exception as e:
        return None, str(e)


def extract_cids_from_logchef_csv(filepath):
    """Extract unique conversation IDs from LogChef export CSV logs"""
    cid_pattern = re.compile(r"([0-9a-f]{11,}-[0-9a-f]{8,})")
    cids = set()
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                log_msg = row.get("log", "")
                matches = cid_pattern.findall(log_msg)
                for match in matches:
                    clean_cid = match.strip('"').strip("'").strip(',').strip(';').strip()
                    if clean_cid:
                        cids.add(clean_cid)
        return list(cids), None
    except Exception as e:
        return None, f"Error extracting CIDs from LogChef CSV: {str(e)}"


@app.route("/api/health", methods=["GET"])
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok"}), 200


@app.route("/api/analyze", methods=["POST"])
def analyze():
    """
    Main analysis endpoint

    Expected:
    - file: CSV, TXT, XLSX file with conversation IDs OR LogChef export CSV
    - time_from: Start time (YYYY-MM-DD HH:MM:SS)
    - time_to: End time (YYYY-MM-DD HH:MM:SS)
    - timezone: Timezone (e.g., Asia/Calcutta)
    """
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Empty filename"}), 400

        allowed_extensions = {'.csv', '.txt', '.xlsx', '.xls'}
        file_ext = '.' + file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            return jsonify({"error": "Unsupported file format. Allowed: CSV, TXT, XLSX"}), 400

        time_from = request.form.get("time_from")
        time_to = request.form.get("time_to")
        timezone = request.form.get("timezone", "Asia/Calcutta")

        if not time_from or not time_to:
            return jsonify({"error": "time_from and time_to are required"}), 400

        try:
            datetime.strptime(time_from, "%Y-%m-%d %H:%M:%S")
            datetime.strptime(time_to, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD HH:MM:SS"}), 400

        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_")
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], timestamp + filename)
        file.save(filepath)

        is_logchef = False
        logchef_csv_path = None

        if file_ext == '.csv':
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    headers = reader.fieldnames or []
                    is_logchef = "pod_name" in headers
            except:
                pass

        if is_logchef:
            print(f"✓ Detected LogChef export format")
            logchef_csv_path = filepath
            cids, error = extract_cids_from_logchef_csv(filepath)
            if error:
                return jsonify({"error": f"Failed to extract CIDs from LogChef CSV: {error}"}), 400
            if not cids:
                return jsonify({"error": "No conversation IDs found in LogChef CSV logs"}), 400
            print(f"✓ Extracted {len(cids)} unique CIDs from LogChef logs")
        else:
            print(f"✓ Reading CIDs from standard file format")
            cids, error = read_cids_from_file(filepath)
            if error:
                return jsonify({"error": f"Failed to read file: {error}"}), 400
            if not cids:
                return jsonify({"error": "No valid conversation IDs found in file"}), 400

        print(f"\n📋 CID Verification:")
        print(f"  Total CIDs extracted: {len(cids)}")
        if cids:
            print(f"  First CID: '{cids[0]}'")
            print(f"  Last CID: '{cids[-1]}'")

        print(f"\n{'='*60}")
        print(f"ANALYSIS CONFIGURATION")
        print(f"{'='*60}")
        print(f"Time window: {time_from} → {time_to}")
        print(f"Timezone: {timezone}")
        if logchef_csv_path:
            print(f"LogChef CSV: {logchef_csv_path}")
        else:
            print(f"Data source: ClickHouse (will fetch logs)")
        print(f"Total CIDs to analyze: {len(cids)}")
        print(f"Sample CIDs: {cids[:3]}{'...' if len(cids) > 3 else ''}")
        print(f"{'='*60}\n")

        try:
            print(f"→ Starting latency analysis...")
            result = run_latency_analysis(
                cids=cids,
                time_from=time_from,
                time_to=time_to,
                timezone=timezone,
                logchef_csv=logchef_csv_path
            )
        except Exception as e:
            print(f"✗ Analysis error: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"error": f"Analysis failed: {str(e)}"}), 500

        print(f"\n✓ Analysis result: {result}")

        if result["status"] != "success":
            return jsonify({"error": result.get("message", "Analysis failed")}), 500

        report_file_path = "LATENCY_REPORT.txt"
        report_filename = f"{timestamp}LATENCY_REPORT.txt"
        report_path = os.path.join(REPORTS_FOLDER, report_filename)
        report_content = None

        print(f"\n📁 Report handling:")
        print(f"  Looking for: {report_file_path}")
        print(f"  CWD: {os.getcwd()}")
        print(f"  Will save to: {report_path}")

        backend_dir = os.path.dirname(os.path.abspath(__file__))
        possible_locations = [
            report_file_path,
            os.path.join(backend_dir, report_file_path),
            os.path.join(os.getcwd(), report_file_path),
        ]

        report_found = False
        for loc in possible_locations:
            if os.path.exists(loc):
                print(f"  ✓ Found report at: {loc}")
                try:
                    with open(loc, "r", encoding="utf-8") as f:
                        report_content = f.read()

                    os.makedirs(REPORTS_FOLDER, exist_ok=True)

                    with open(report_path, "w", encoding="utf-8") as f:
                        f.write(report_content)

                    os.remove(loc)
                    print(f"  ✓ Report saved to: {report_path}")
                    report_found = True
                    break
                except Exception as e:
                    print(f"  ✗ Error processing report at {loc}: {e}")
                    import traceback
                    traceback.print_exc()

        if not report_found:
            print(f"  ✗ Report file not found in any location")
            return jsonify({"error": "Report file not generated"}), 500

        if not report_content:
            print(f"  ✗ Report file is empty")
            return jsonify({"error": "Report file is empty"}), 500

        return jsonify({
            "status": "success",
            "message": f"Analysis complete for {len(cids)} conversations",
            "download_url": f"/api/download/{report_filename}",
            "report_content": report_content,
            "cids_processed": len(cids)
        }), 200

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/test-cid-extraction", methods=["POST"])
def test_cid_extraction():
    """Test endpoint to verify file parsing and preview CIDs"""
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], "test_" + filename)
        file.save(filepath)

        cids, error = read_cids_from_file(filepath)
        os.remove(filepath)

        if error:
            return jsonify({"error": error}), 400

        return jsonify({
            "status": "success",
            "cids_found": len(cids),
            "sample_cids": cids[:5] if len(cids) > 5 else cids
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download/<filename>", methods=["GET"])
def download(filename):
    """Download generated report"""
    try:
        if ".." in filename or "/" in filename or "\\" in filename:
            return jsonify({"error": "Invalid filename"}), 400

        filepath = os.path.join(REPORTS_FOLDER, filename)

        if not os.path.exists(filepath):
            print(f"File not found: {filepath}")
            return jsonify({"error": f"File not found: {filename}"}), 404

        if not os.path.abspath(filepath).startswith(os.path.abspath(REPORTS_FOLDER)):
            return jsonify({"error": "Invalid file path"}), 403

        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype="text/plain"
        )
    except Exception as e:
        print(f"Download error: {str(e)}")
        return jsonify({"error": f"Download failed: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
